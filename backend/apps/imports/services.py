from __future__ import annotations

import csv
from zipfile import BadZipFile
from dataclasses import dataclass
from io import StringIO

from django.db import transaction
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook

from apps.catalog.models import Brand, Category, Manufacturer, Part, PartNumber, PriceOffer, Supplier
from apps.catalog.services import build_part_search_text, normalize_part_number, stable_slug
from .models import PriceImport


class PriceImportParseError(ValueError):
    pass


@dataclass
class PriceImportRow:
    name: str
    article: str
    oem: str = ""
    brand: str = "Zemazap"
    model: str = "уточнить"
    category: str = "Импорт"
    manufacturer: str = "уточнить"
    price: int = 0
    availability: str = "уточнить"
    stock: str = "уточнить"
    delivery: str = "Срок и наличие подтверждает менеджер"
    condition: str = "новая"
    photo_kind: str = "illustrative"
    warranty_terms: str = ""
    return_terms: str = ""
    marking_required: bool = False
    marking_status: str = "not_required"
    marking_category: str = ""


COLUMN_ALIASES = {
    "название": "name",
    "наименование": "name",
    "товар": "name",
    "name": "name",
    "артикул": "article",
    "article": "article",
    "sku": "article",
    "oem": "oem",
    "номер": "oem",
    "номер детали": "oem",
    "марка": "brand",
    "brand": "brand",
    "модель": "model",
    "model": "model",
    "категория": "category",
    "category": "category",
    "производитель": "manufacturer",
    "manufacturer": "manufacturer",
    "бренд детали": "manufacturer",
    "цена": "price",
    "price": "price",
    "стоимость": "price",
    "наличие": "availability",
    "availability": "availability",
    "склад": "stock",
    "stock": "stock",
    "срок": "delivery",
    "delivery": "delivery",
    "срок поставки": "delivery",
    "состояние": "condition",
    "condition": "condition",
    "тип фото": "photo_kind",
    "photo kind": "photo_kind",
    "гарантия": "warranty_terms",
    "warranty": "warranty_terms",
    "условия возврата": "return_terms",
    "return terms": "return_terms",
    "маркировка обязательна": "marking_required",
    "marking required": "marking_required",
    "статус маркировки": "marking_status",
    "marking status": "marking_status",
    "категория маркировки": "marking_category",
    "marking category": "marking_category",
}
AVAILABILITY_VALUES = {"в наличии", "1-3 дня", "под заказ", "уточнить"}
CONDITION_VALUES = {choice for choice, _label in Part.Condition.choices}
PHOTO_KIND_VALUES = {choice for choice, _label in Part.PhotoKind.choices}
MARKING_STATUS_VALUES = {choice for choice, _label in Part.MarkingStatus.choices}
CSV_ENCODINGS = ("utf-8-sig", "cp1251")


def normalize_header(value: str) -> str:
    return " ".join(value.strip().lower().split())


def parse_price(value: object) -> int:
    if isinstance(value, int | float):
        return max(0, round(value))
    cleaned = "".join(char for char in str(value) if char.isdigit() or char in ",.")
    try:
        return max(0, round(float(cleaned.replace(",", "."))))
    except ValueError:
        return 0


def decode_csv_content(content: bytes) -> str:
    for encoding in CSV_ENCODINGS:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise PriceImportParseError("Не удалось прочитать CSV: поддерживаются UTF-8, UTF-8 BOM и Windows-1251.")


def detect_csv_delimiter(text: str) -> str:
    sample = text[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        return dialect.delimiter
    except csv.Error:
        first_lines = [line for line in text.splitlines()[:10] if line.strip()]
        semicolons = sum(line.count(";") for line in first_lines)
        commas = sum(line.count(",") for line in first_lines)
        return "," if commas > semicolons else ";"


def rows_from_csv(content: bytes) -> list[list[str]]:
    text = decode_csv_content(content)
    delimiter = detect_csv_delimiter(text)
    return [row for row in csv.reader(StringIO(text), delimiter=delimiter) if any(cell.strip() for cell in row)]


def rows_from_xlsx(content: bytes) -> list[list[str]]:
    workbook = load_workbook(filename=StringIO(content.decode("latin1")), read_only=True, data_only=True)
    sheet = workbook.active
    return [[str(cell) if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]


def parse_table(table: list[list[object]]) -> list[PriceImportRow]:
    if not table:
        return []
    headers = [COLUMN_ALIASES.get(normalize_header(str(value))) for value in table[0]]
    parsed: list[PriceImportRow] = []

    for raw_row in table[1:]:
        item: dict[str, object] = {}
        for index, value in enumerate(raw_row):
            key = headers[index] if index < len(headers) else None
            if not key or value in ("", None):
                continue
            if key == "price":
                item[key] = parse_price(value)
            elif key == "availability":
                availability = str(value).strip().lower()
                item[key] = availability if availability in AVAILABILITY_VALUES else "уточнить"
            elif key == "condition":
                condition = str(value).strip().lower()
                item[key] = condition if condition in CONDITION_VALUES else Part.Condition.NEW
            elif key == "photo_kind":
                photo_kind = str(value).strip().lower()
                item[key] = photo_kind if photo_kind in PHOTO_KIND_VALUES else Part.PhotoKind.ILLUSTRATIVE
            elif key == "marking_status":
                marking_status = str(value).strip().lower()
                item[key] = marking_status if marking_status in MARKING_STATUS_VALUES else Part.MarkingStatus.REQUIRES_REVIEW
            elif key == "marking_required":
                item[key] = str(value).strip().lower() in {"1", "true", "yes", "да", "обязательно"}
            else:
                item[key] = str(value).strip()
        if item.get("name") and item.get("article"):
            parsed.append(PriceImportRow(**item))
    return parsed


def parse_price_import_file(filename: str, content: bytes) -> list[PriceImportRow]:
    if filename.lower().endswith(".xlsx"):
        from io import BytesIO

        try:
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            table = [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise PriceImportParseError("Не удалось прочитать XLSX: файл поврежден или имеет неверный формат.") from exc
    else:
        table = rows_from_csv(content)
    rows = parse_table(table)
    if not rows:
        raise PriceImportParseError("В файле не найдены строки с обязательными колонками: название и артикул.")
    return rows


def get_or_create_brand(value: str) -> Brand:
    slug = stable_slug(value, "zemazap")
    return Brand.objects.get_or_create(slug=slug, defaults={"name": value, "country": "уточнить", "sort_order": 999})[0]


def get_or_create_category(value: str) -> Category:
    slug = stable_slug(value, "import")
    return Category.objects.get_or_create(
        slug=slug,
        defaults={"name": value, "description": "Импортированная категория, описание нужно заполнить.", "sort_order": 999},
    )[0]


@transaction.atomic
def import_price_rows(filename: str, file_kind: str, rows: list[PriceImportRow]) -> PriceImport:
    supplier, _created = Supplier.objects.get_or_create(
        name="Импортированный прайс",
        defaults={"kind": "import", "contact_note": "Источник импортированных CSV/XLSX-позиций"},
    )
    imported_rows = 0
    skipped_rows = 0

    for row in rows:
        if not row.name or not row.article:
            skipped_rows += 1
            continue
        brand = get_or_create_brand(row.brand)
        category = get_or_create_category(row.category)
        manufacturer, _created = Manufacturer.objects.get_or_create(name=row.manufacturer or "уточнить")
        part_slug = stable_slug(f"{row.article}-{row.name}", f"import-{imported_rows + 1}", max_length=150)
        legacy_id = f"import:{normalize_part_number(row.article) or part_slug}"
        oem = row.oem or row.article
        search_text = build_part_search_text(
            name=row.name,
            oem=oem,
            article=row.article,
            manufacturer=manufacturer.name,
            category=category.name,
            brand=brand.name,
            model=row.model,
            condition=row.condition,
            quality="заводской аналог",
            analogs=[],
            compatibility=[],
            specs={},
        )
        part, _created = Part.objects.update_or_create(
            legacy_id=legacy_id,
            defaults={
                "slug": part_slug,
                "name": row.name,
                "category": category,
                "brand": brand,
                "model_name": row.model,
                "manufacturer": manufacturer,
                "condition": row.condition,
                "photo_kind": row.photo_kind,
                "warranty_terms": row.warranty_terms,
                "return_terms": row.return_terms,
                "marking_required": row.marking_required,
                "marking_status": row.marking_status,
                "marking_category": row.marking_category,
                "quality": "заводской аналог",
                "description": "Импортированная позиция. Описание, применимость и гарантию нужно уточнить.",
                "primary_oem": oem,
                "primary_article": row.article,
                "search_text": search_text,
                "is_active": True,
                "sort_order": 999,
            },
        )
        part.numbers.all().delete()
        PartNumber.objects.create(part=part, kind=PartNumber.Kind.ARTICLE, value=row.article, normalized_value=normalize_part_number(row.article), sort_order=0)
        PartNumber.objects.create(part=part, kind=PartNumber.Kind.OEM, value=oem, normalized_value=normalize_part_number(oem), sort_order=1)
        part.price_offers.filter(is_primary=True).delete()
        PriceOffer.objects.create(
            part=part,
            supplier=supplier,
            price_rub=row.price,
            availability=row.availability,
            delivery=row.delivery,
            stock=row.stock,
            is_primary=True,
        )
        imported_rows += 1

    return PriceImport.objects.create(
        filename=filename,
        file_kind=file_kind,
        imported_rows=imported_rows,
        skipped_rows=skipped_rows,
    )
